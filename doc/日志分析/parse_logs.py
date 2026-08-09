#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
巡线 log 解析器 -> Excel。

【最简单用法】双击 "解析log.bat"（或直接双击本 .py）：
  弹出文件选择框 -> 选一个 log 的 .txt -> 在同目录生成同名 .xlsx -> 问你是否打开。

【命令行用法】
  python parse_logs.py log19.txt            # 解析单个文件 -> log19.xlsx
  python parse_logs.py --all                # 解析脚本目录下所有 *.txt -> logs_parsed.xlsx
  python parse_logs.py --input x.txt --output out.xlsx

解析对象(见 log 里的 "log line format"):
  E<dt> <patHex> <modeCode> <L> <R>[ <err>]   (H=heartbeat 代替 E)
    E行 dt = 距上一次档位切换的ms; H行 dt = 距上一条记录行的ms
    patHex = 2位十六进制位图 (bit7=CH8 .. bit0=CH1, 1=白)
    modeCode: 0=STRAIGHT 1/2=unused 3=MEDIUM_L 4=MEDIUM_R 5=SHARP_L 6=SHARP_R
              7=HAIRPIN_L 8=HAIRPIN_R 9=CROSS A=LOST_L B=LOST_R C=LOST_STOP P=PID
    L/R = 实际下发PWM; err = PID误差(仅PID模式)

每个 log 文件 -> 一个 sheet，每条 E/H 行 -> 一行，解码出各字段。
标记行(TRACK_ON / PARAMS / TRACK_OFF)用于给每行补上 section 上下文(算法id/name/speed等)。
需要依赖: openpyxl  (pip install openpyxl)
"""

import argparse
import os
import re
import sys

MODE_NAMES = {
    '0': 'STRAIGHT', '1': 'LEFT(mild,unused)', '2': 'RIGHT(mild,unused)',
    '3': 'MEDIUM_L', '4': 'MEDIUM_R', '5': 'SHARP_L', '6': 'SHARP_R',
    '7': 'HAIRPIN_L', '8': 'HAIRPIN_R', '9': 'CROSS',
    'A': 'LOST_L', 'B': 'LOST_R', 'C': 'LOST_STOP', 'P': 'PID',
}

# 物理左->右: CH8 CH7 CH6 CH5 | CH4 CH3 CH2 CH1  (CH1最右)
# 位置(正=偏右, 与PID err符号一致): CH8=-3.5 ... CH1=+3.5
CH_POS = {8: -3.5, 7: -2.5, 6: -1.5, 5: -0.5, 4: 0.5, 3: 1.5, 2: 2.5, 1: 3.5}

# E/H 数据行. modeCode 允许 0-9,A-C,P (大小写皆可)
LINE_RE = re.compile(
    r'^([EH])(\d+)\s+([0-9A-Fa-f]{2})\s+([0-9A-CPa-cp])\s+(-?\d+)\s+(-?\d+)'
    r'(?:\s+([+-]?\d+(?:\.\d+)?))?\s*$'
)
TRACK_ON_RE = re.compile(r'TRACK_ON\s+t=(\d+)')
TRACK_OFF_RE = re.compile(r'TRACK_OFF\b')
PARAMS_RE = re.compile(r'^>>>\s*PARAMS\b(.*)$')

COLUMNS = [
    'line_no', 'sec', 'ev', 'dt_ms', 't_ms',
    'patHex',
    # 物理左->右: CH8 CH7 CH6 CH5 | CH4 CH3 CH2 CH1 (与车上传感器方向一致)
    'CH8', 'CH7', 'CH6', 'CH5', 'CH4', 'CH3', 'CH2', 'CH1',
    'white_chs', 'n_white', 'pos_LtoR',
    'mode', 'mode_name', 'L', 'R', 'err',
    'raw',
]

# 参数在顶部区块显示的优先顺序(其余未列出的键按字母序追加在后)
PARAM_KEY_ORDER = [
    'algoid', 'name', 'algo', 'speed', 'slew_rate',
    'turn_ratio', 'medium_ratio', 'sharp_ratio', 'xsharp_ratio',
    'medspeed', 'shpspeed', 'hpspeed', 'minpwm',
    'pid_kp', 'pid_ki', 'pid_kd',
    'thresh', 'white', 'black',
]


def parse_params(text):
    d = {}
    for tok in text.strip().split():
        if '=' in tok:
            k, v = tok.split('=', 1)
            d[k] = v
    return d


def decode_pattern(pat_int):
    chvals = {}
    white = []
    for ch in range(1, 9):
        bit = (pat_int >> (ch - 1)) & 1
        chvals[ch] = bit
        if bit:
            white.append(ch)
    n = len(white)
    pos = round(sum(CH_POS[c] for c in white) / n, 3) if n else None
    # 按物理左->右列出(CH8最左 .. CH1最右)，与列顺序一致
    white_list = ','.join('CH%d' % c for c in sorted(white, reverse=True))
    return chvals, white_list, n, pos


def parse_file(path):
    """返回 (rows, sections)。sections=[{'sec','track_on','params'{}}]，供顶部参数区块用。"""
    rows = []
    sections = []
    cur = None            # 当前 section dict
    sec = -1
    last_switch = 0.0
    last_line = 0.0
    with open(path, 'r', encoding='utf-8', errors='replace') as f:
        for i, line in enumerate(f, 1):
            s = line.rstrip('\r\n')

            mon = TRACK_ON_RE.search(s)
            if mon:
                sec += 1
                cur = {'sec': sec, 'track_on': mon.group(1), 'params': {}}
                sections.append(cur)
                last_switch = 0.0
                last_line = 0.0
                continue
            m = PARAMS_RE.match(s)
            if m:
                if cur is None:   # 没见 TRACK_ON 就先来了 PARAMS，兜底建一段
                    sec += 1
                    cur = {'sec': sec, 'track_on': '', 'params': {}}
                    sections.append(cur)
                cur['params'] = parse_params(m.group(1))
                continue
            if TRACK_OFF_RE.search(s) and s.lstrip().startswith('>>>'):
                continue

            m = LINE_RE.match(s.strip())
            if not m:
                continue

            ev, dt, pathex, mode, L, R, err = m.groups()
            dt = int(dt)
            mode = mode.upper()
            pat_int = int(pathex, 16)

            if ev == 'E':
                t = last_switch + dt
                last_switch = t
                last_line = t
            else:
                t = last_line + dt
                last_line = t

            chvals, white_list, n_white, pos = decode_pattern(pat_int)

            rows.append({
                'line_no': i, 'sec': sec, 'ev': ev, 'dt_ms': dt, 't_ms': round(t, 1),
                'patHex': pathex.upper(),
                'CH1': chvals[1], 'CH2': chvals[2], 'CH3': chvals[3], 'CH4': chvals[4],
                'CH5': chvals[5], 'CH6': chvals[6], 'CH7': chvals[7], 'CH8': chvals[8],
                'white_chs': white_list, 'n_white': n_white, 'pos_LtoR': pos,
                'mode': mode, 'mode_name': MODE_NAMES.get(mode, '?'),
                'L': int(L), 'R': int(R), 'err': float(err) if err is not None else None,
                'raw': s.strip(),
            })
    return rows, sections


def _sheet_name(base, used):
    name = base[:31]
    n = 1
    while name in used:
        suffix = '_%d' % n
        name = base[:31 - len(suffix)] + suffix
        n += 1
    used.add(name)
    return name


def _ordered_param_keys(sections):
    allk = set()
    for s in sections:
        allk.update(s['params'].keys())
    ordered = [k for k in PARAM_KEY_ORDER if k in allk]
    ordered += sorted(k for k in allk if k not in PARAM_KEY_ORDER)
    return ordered


def write_workbook(sheets, out):
    """sheets = {sheet名: (rows, sections)}。写成 xlsx。openpyxl 缺失会抛 ImportError。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font as _Font
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    wb.remove(wb.active)
    used = set()
    widths = {'white_chs': 14, 'mode_name': 18, 'raw': 60, 'patHex': 8}
    bold = _Font(bold=True)

    for base, (rows, sections) in sheets.items():
        ws = wb.create_sheet(_sheet_name(base, used))

        # ---- 顶部:参数区块(表头上方),每个 TRACK_ON 段一行 ----
        c = ws.cell(row=1, column=1, value='PARAMS (每个 TRACK_ON 段一行, 全部参数):')
        c.font = bold
        keys = _ordered_param_keys(sections)
        # 参数小表: 表头 [sec, track_on, <各参数键>]
        phdr = ['sec', 'track_on'] + keys
        ws.append(phdr)
        for cell in ws[ws.max_row]:
            cell.font = bold
        for s in sections:
            ws.append([s['sec'], s['track_on']] + [s['params'].get(k, '') for k in keys])
        ws.append([])  # 空行分隔

        # ---- 数据表头 + 数据 ----
        ws.append(COLUMNS)
        header_row = ws.max_row
        for cell in ws[header_row]:
            cell.font = bold
        for r in rows:
            ws.append([r.get(col) for col in COLUMNS])
        ws.freeze_panes = 'A%d' % (header_row + 1)   # 冻结到数据表头下一行
        for idx, col in enumerate(COLUMNS, 1):
            ws.column_dimensions[get_column_letter(idx)].width = widths.get(col, 8)
    wb.save(out)


def run_gui():
    """双击运行:弹文件框选一个 log txt -> 同名 xlsx。结果用弹窗提示。"""
    try:
        import tkinter as tk
        from tkinter import filedialog, messagebox
    except Exception:
        # 没有图形界面就退回命令行提示
        print('无法启动图形界面，请用命令行: python parse_logs.py <文件.txt>')
        return
    root = tk.Tk()
    root.withdraw()
    start_dir = os.path.dirname(os.path.abspath(__file__))
    path = filedialog.askopenfilename(
        title='选择一个 log 的 .txt 文件',
        initialdir=start_dir,
        filetypes=[('日志文本', '*.txt'), ('所有文件', '*.*')],
    )
    if not path:
        return
    try:
        rows, sections = parse_file(path)
    except Exception as e:
        messagebox.showerror('解析失败', str(e))
        return
    if not rows:
        messagebox.showwarning('无数据', '这个文件里没解析到 E/H 数据行。')
        return
    out = os.path.splitext(path)[0] + '.xlsx'
    base = os.path.splitext(os.path.basename(path))[0]
    try:
        write_workbook({base: (rows, sections)}, out)
    except ImportError:
        messagebox.showerror('缺少依赖', '需要 openpyxl。\n请先运行:\n    python -m pip install openpyxl')
        return
    except PermissionError:
        messagebox.showerror('保存失败', '无法写入(文件可能正被 Excel 打开):\n%s' % out)
        return
    except Exception as e:
        messagebox.showerror('保存失败', str(e))
        return
    if messagebox.askyesno('完成', '已解析 %d 行 ->\n%s\n\n现在打开它吗?' % (len(rows), out)):
        try:
            os.startfile(out)  # 仅 Windows
        except Exception:
            pass


def main():
    ap = argparse.ArgumentParser(description='巡线 log -> Excel 解析器')
    ap.add_argument('input', nargs='?', default=None, help='单个 log 文件(.txt)')
    ap.add_argument('--input', dest='input_opt', default=None, help='同上(可选写法)')
    ap.add_argument('--all', action='store_true', help='解析脚本目录下所有 *.txt')
    ap.add_argument('--output', '-o', default=None, help='输出 xlsx 路径')
    args = ap.parse_args()

    script_dir = os.path.dirname(os.path.abspath(__file__))
    inp = args.input_opt or args.input

    if not args.all and not inp:
        run_gui()   # 无参数 = 双击 = 图形选择
        return

    if args.all:
        files = [os.path.join(script_dir, fn) for fn in sorted(os.listdir(script_dir))
                 if fn.lower().endswith('.txt')]
        out = args.output or os.path.join(script_dir, 'logs_parsed.xlsx')
    else:
        files = [inp]
        out = args.output or (os.path.splitext(inp)[0] + '.xlsx')

    sheets = {}
    total = 0
    for path in files:
        try:
            rows, sections = parse_file(path)
        except Exception as e:
            print('跳过 %s: %s' % (path, e), file=sys.stderr)
            continue
        if not rows:
            continue
        sheets[os.path.splitext(os.path.basename(path))[0]] = (rows, sections)
        total += len(rows)
        print('  %-24s %d 行' % (os.path.basename(path), len(rows)))

    if not sheets:
        print('没有可解析的数据行。', file=sys.stderr)
        sys.exit(1)
    try:
        write_workbook(sheets, out)
    except ImportError:
        print('需要 openpyxl: pip install openpyxl', file=sys.stderr)
        sys.exit(1)
    print('完成: %d 个文件, 共 %d 行 -> %s' % (len(sheets), total, out))


if __name__ == '__main__':
    main()
